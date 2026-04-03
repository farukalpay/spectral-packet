from __future__ import annotations

import csv
from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import numpy as np


_DELIMITERS = {
    ".csv": ",",
    ".tsv": "\t",
}


def parquet_support_is_available() -> bool:
    try:
        import pyarrow  # noqa: F401
        import pyarrow.parquet  # noqa: F401
    except ModuleNotFoundError:
        return False
    return True


def xlsx_tabular_support_is_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ModuleNotFoundError:
        return False
    return True


def supported_tabular_formats() -> dict[str, bool]:
    return {
        "csv": True,
        "tsv": True,
        "json": True,
        "parquet": parquet_support_is_available(),
        "xlsx": xlsx_tabular_support_is_available(),
    }


def _supported_tabular_format_labels() -> str:
    supported = supported_tabular_formats()
    available = [f".{name}" for name, enabled in supported.items() if enabled]
    unavailable = [f".{name}" for name, enabled in supported.items() if not enabled]
    if not unavailable:
        return ", ".join(available)
    return f"{', '.join(available)} (optional: {', '.join(unavailable)})"


def _is_missing(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value == "")


def _infer_text_dtype(values: Sequence[Any]) -> str:
    if all(isinstance(value, str) or _is_missing(value) for value in values):
        return "string"
    return "object"


def _coerce_numeric(values: Sequence[Any]) -> np.ndarray | None:
    converted: list[float] = []
    integer_like = True
    has_missing = False
    for value in values:
        if _is_missing(value):
            has_missing = True
            converted.append(np.nan)
            continue
        if isinstance(value, bool):
            return None
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        if not float(number).is_integer():
            integer_like = False
        converted.append(number)
    if not converted:
        return np.asarray([], dtype=np.float64)
    if integer_like and not has_missing:
        return np.asarray(converted, dtype=np.int64)
    return np.asarray(converted, dtype=np.float64)


def _coerce_column(values: Sequence[Any]) -> np.ndarray:
    if not values:
        return np.asarray([], dtype=np.float64)

    numeric = _coerce_numeric(values)
    if numeric is not None:
        return numeric

    if all(isinstance(value, bool) or _is_missing(value) for value in values):
        return np.asarray(
            [None if _is_missing(value) else bool(value) for value in values],
            dtype=object,
        )

    text_dtype = _infer_text_dtype(values)
    if text_dtype == "string":
        return np.asarray(
            ["" if value is None else str(value) for value in values],
            dtype=str,
        )
    return np.asarray(list(values), dtype=object)


def _dtype_name(values: np.ndarray) -> str:
    if values.dtype.kind in {"i", "u"}:
        return "int64"
    if values.dtype.kind == "f":
        return "float64"
    if values.dtype.kind == "b":
        return "bool"
    if values.dtype.kind in {"U", "S"}:
        return "string"
    return "object"


def _column_missing_count(values: np.ndarray) -> int:
    if values.dtype.kind == "f":
        return int(np.isnan(values).sum())
    if values.dtype.kind in {"U", "S"}:
        return int(sum(1 for value in values.tolist() if value == ""))
    return int(sum(1 for value in values.tolist() if value is None))


def _column_non_finite_count(values: np.ndarray) -> int:
    if values.dtype.kind in {"i", "u"}:
        return 0
    if values.dtype.kind == "f":
        finite = np.isfinite(values)
        missing = np.isnan(values)
        return int((~finite & ~missing).sum())
    return 0


@dataclass(frozen=True, slots=True)
class TabularSource:
    kind: str
    location: str | None = None
    description: str | None = None


@dataclass(frozen=True, slots=True)
class TabularColumn:
    name: str
    dtype: str
    nullable: bool


@dataclass(frozen=True, slots=True)
class TabularSchema:
    columns: tuple[TabularColumn, ...]
    row_count: int

    @property
    def column_names(self) -> tuple[str, ...]:
        return tuple(column.name for column in self.columns)


@dataclass(frozen=True, slots=True)
class JoinSpec:
    left_keys: tuple[str, ...]
    right_keys: tuple[str, ...]
    how: str = "inner"

    def __post_init__(self) -> None:
        if not self.left_keys or not self.right_keys:
            raise ValueError("join keys must not be empty")
        if len(self.left_keys) != len(self.right_keys):
            raise ValueError("left_keys and right_keys must have the same length")
        if self.how not in {"inner", "left"}:
            raise ValueError("how must be 'inner' or 'left'")


@dataclass(frozen=True, slots=True)
class TabularValidationReport:
    row_count: int
    column_count: int
    missing_value_counts: dict[str, int]
    non_finite_numeric_counts: dict[str, int]


@dataclass(frozen=True, slots=True)
class TabularDataset:
    columns: dict[str, np.ndarray]
    source: TabularSource | None = None

    def __post_init__(self) -> None:
        if not self.columns:
            raise ValueError("columns must not be empty")
        normalized: dict[str, np.ndarray] = {}
        row_count: int | None = None
        for name, values in self.columns.items():
            column_name = str(name)
            if not column_name:
                raise ValueError("column names must not be empty")
            array = np.asarray(values)
            if array.ndim != 1:
                raise ValueError(f"column '{column_name}' must be one-dimensional")
            if row_count is None:
                row_count = int(array.shape[0])
            elif int(array.shape[0]) != row_count:
                raise ValueError("all columns must have the same number of rows")
            normalized[column_name] = array
        object.__setattr__(self, "columns", normalized)

    @classmethod
    def from_rows(
        cls,
        rows: Sequence[Mapping[str, Any]],
        *,
        source: TabularSource | None = None,
    ) -> TabularDataset:
        if not rows:
            raise ValueError("rows must not be empty")
        column_names = list(rows[0].keys())
        if not column_names:
            raise ValueError("rows must contain at least one column")
        for row in rows:
            if set(row.keys()) != set(column_names):
                raise ValueError("all rows must contain the same columns")
        columns = {
            name: _coerce_column([row.get(name) for row in rows])
            for name in column_names
        }
        return cls(columns=columns, source=source)

    @property
    def row_count(self) -> int:
        return int(next(iter(self.columns.values())).shape[0])

    @property
    def column_names(self) -> tuple[str, ...]:
        return tuple(self.columns.keys())

    @property
    def schema(self) -> TabularSchema:
        return TabularSchema(
            columns=tuple(
                TabularColumn(
                    name=name,
                    dtype=_dtype_name(values),
                    nullable=_column_missing_count(values) > 0,
                )
                for name, values in self.columns.items()
            ),
            row_count=self.row_count,
        )

    def validation_report(self) -> TabularValidationReport:
        return TabularValidationReport(
            row_count=self.row_count,
            column_count=len(self.columns),
            missing_value_counts={
                name: _column_missing_count(values)
                for name, values in self.columns.items()
            },
            non_finite_numeric_counts={
                name: _column_non_finite_count(values)
                for name, values in self.columns.items()
                if _column_non_finite_count(values) > 0
            },
        )

    def to_rows(self, *, limit: int | None = None) -> list[dict[str, Any]]:
        row_limit = self.row_count if limit is None else min(limit, self.row_count)
        rows: list[dict[str, Any]] = []
        for index in range(row_limit):
            row: dict[str, Any] = {}
            for name, values in self.columns.items():
                value = values[index]
                if isinstance(value, np.generic):
                    row[name] = value.item()
                else:
                    row[name] = value
            rows.append(row)
        return rows

    def select_columns(self, columns: Sequence[str]) -> TabularDataset:
        selected = [str(name) for name in columns]
        missing = [name for name in selected if name not in self.columns]
        if missing:
            raise ValueError(f"unknown columns: {', '.join(missing)}")
        return TabularDataset(
            columns={name: self.columns[name].copy() for name in selected},
            source=self.source,
        )

    def rename_columns(self, mapping: Mapping[str, str]) -> TabularDataset:
        renamed: dict[str, np.ndarray] = {}
        for name, values in self.columns.items():
            target = str(mapping.get(name, name))
            if not target:
                raise ValueError("renamed column names must not be empty")
            if target in renamed:
                raise ValueError(f"duplicate column name after rename: {target}")
            renamed[target] = values.copy()
        return TabularDataset(columns=renamed, source=self.source)

    def with_column(self, name: str, values: Sequence[Any]) -> TabularDataset:
        array = _coerce_column(list(values))
        if array.shape[0] != self.row_count:
            raise ValueError("new column must have the same row count as the dataset")
        updated = {column_name: column_values.copy() for column_name, column_values in self.columns.items()}
        updated[str(name)] = array
        return TabularDataset(columns=updated, source=self.source)

    def filter_rows(self, mask: Sequence[bool]) -> TabularDataset:
        mask_array = np.asarray(mask, dtype=bool)
        if mask_array.ndim != 1 or mask_array.shape[0] != self.row_count:
            raise ValueError("mask must be one-dimensional and match the row count")
        return TabularDataset(
            columns={name: values[mask_array] for name, values in self.columns.items()},
            source=self.source,
        )

    def sort_by(self, column: str, *, descending: bool = False) -> TabularDataset:
        if column not in self.columns:
            raise ValueError(f"unknown column: {column}")
        order = np.argsort(self.columns[column], kind="stable")
        if descending:
            order = order[::-1]
        return TabularDataset(
            columns={name: values[order] for name, values in self.columns.items()},
            source=self.source,
        )

    def join(
        self,
        other: TabularDataset,
        *,
        spec: JoinSpec,
        right_suffix: str = "_right",
    ) -> TabularDataset:
        for name in spec.left_keys:
            if name not in self.columns:
                raise ValueError(f"unknown left join column: {name}")
        for name in spec.right_keys:
            if name not in other.columns:
                raise ValueError(f"unknown right join column: {name}")

        right_index: dict[tuple[Any, ...], list[int]] = {}
        for right_row_index in range(other.row_count):
            key = tuple(other.columns[name][right_row_index].item() if isinstance(other.columns[name][right_row_index], np.generic) else other.columns[name][right_row_index] for name in spec.right_keys)
            right_index.setdefault(key, []).append(right_row_index)

        joined_rows: list[dict[str, Any]] = []
        for left_row_index in range(self.row_count):
            left_key = tuple(self.columns[name][left_row_index].item() if isinstance(self.columns[name][left_row_index], np.generic) else self.columns[name][left_row_index] for name in spec.left_keys)
            matches = right_index.get(left_key, [])
            if not matches and spec.how == "left":
                matches = [None]
            for right_row_index in matches:
                if right_row_index is None and spec.how != "left":
                    continue
                row: dict[str, Any] = {}
                for name, values in self.columns.items():
                    value = values[left_row_index]
                    row[name] = value.item() if isinstance(value, np.generic) else value
                for name, values in other.columns.items():
                    if name in spec.right_keys:
                        continue
                    target_name = name if name not in row else f"{name}{right_suffix}"
                    if right_row_index is None:
                        row[target_name] = None
                    else:
                        value = values[right_row_index]
                        row[target_name] = value.item() if isinstance(value, np.generic) else value
                joined_rows.append(row)

        if not joined_rows:
            columns = {
                name: np.asarray([], dtype=values.dtype)
                for name, values in self.columns.items()
            }
            for name, values in other.columns.items():
                if name in spec.right_keys:
                    continue
                target_name = name if name not in columns else f"{name}{right_suffix}"
                columns[target_name] = np.asarray([], dtype=values.dtype)
            return TabularDataset(columns=columns, source=self.source)
        return TabularDataset.from_rows(joined_rows, source=self.source)

    def numeric_matrix(self, columns: Sequence[str]) -> np.ndarray:
        selected = []
        for name in columns:
            if name not in self.columns:
                raise ValueError(f"unknown column: {name}")
            values = self.columns[name]
            if values.dtype.kind not in {"i", "u", "f"}:
                raise ValueError(f"column '{name}' is not numeric")
            selected.append(values.astype(np.float64, copy=False))
        return np.column_stack(selected)


def _load_tabular_dataset_delimited(
    path: str | Path,
    *,
    delimiter: str,
) -> TabularDataset:
    input_path = Path(path)
    with input_path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if reader.fieldnames is None:
            raise ValueError("tabular file must contain a header row")
        rows = [row for row in reader if row is not None and any(value != "" for value in row.values())]
    if not rows:
        raise ValueError("tabular file contains no data rows")
    return TabularDataset.from_rows(
        rows,
        source=TabularSource(kind="file", location=str(input_path)),
    )


def load_tabular_dataset_csv(path: str | Path) -> TabularDataset:
    return _load_tabular_dataset_delimited(path, delimiter=",")


def load_tabular_dataset_tsv(path: str | Path) -> TabularDataset:
    return _load_tabular_dataset_delimited(path, delimiter="\t")


def load_tabular_dataset_json(path: str | Path) -> TabularDataset:
    input_path = Path(path)
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        rows = payload.get("rows")
    else:
        rows = payload
    if not isinstance(rows, list) or not rows or not all(isinstance(row, Mapping) for row in rows):
        raise ValueError("JSON tabular data must be a non-empty array of objects or an object with a 'rows' array")
    return TabularDataset.from_rows(
        rows,
        source=TabularSource(kind="file", location=str(input_path)),
    )


def load_tabular_dataset_parquet(path: str | Path) -> TabularDataset:
    try:
        import pyarrow.parquet as pq
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Parquet support requires pyarrow. Install pyarrow in the active environment."
        ) from exc

    input_path = Path(path)
    table = pq.read_table(input_path)
    columns = {
        str(name): np.asarray(table.column(name).to_pylist())
        for name in table.column_names
    }
    return TabularDataset(
        columns=columns,
        source=TabularSource(kind="file", location=str(input_path)),
    )


def load_tabular_dataset_xlsx(path: str | Path) -> TabularDataset:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "XLSX tabular support requires openpyxl. Install the 'files' extra."
        ) from exc

    input_path = Path(path)
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("XLSX file is empty")
        header = ["" if value is None else str(value) for value in rows[0]]
        if not any(header):
            raise ValueError("XLSX file must contain a header row")
        row_objects: list[dict[str, Any]] = []
        for row in rows[1:]:
            values = list(row)
            if not any(value is not None and value != "" for value in values):
                continue
            if len(values) < len(header):
                values.extend([None] * (len(header) - len(values)))
            row_objects.append(
                {
                    column_name: values[index]
                    for index, column_name in enumerate(header)
                }
            )
        if not row_objects:
            raise ValueError("XLSX file contains no data rows")
        return TabularDataset.from_rows(
            row_objects,
            source=TabularSource(kind="file", location=str(input_path)),
        )
    finally:
        workbook.close()


def load_tabular_dataset(path: str | Path) -> TabularDataset:
    input_path = Path(path)
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return load_tabular_dataset_csv(input_path)
    if suffix == ".tsv":
        return load_tabular_dataset_tsv(input_path)
    if suffix == ".json":
        return load_tabular_dataset_json(input_path)
    if suffix == ".parquet":
        return load_tabular_dataset_parquet(input_path)
    if suffix == ".xlsx":
        return load_tabular_dataset_xlsx(input_path)
    supported = _supported_tabular_format_labels()
    raise ValueError(f"unsupported tabular format '{suffix or '<none>'}'; supported formats: {supported}")


def _save_tabular_dataset_delimited(
    dataset: TabularDataset,
    path: str | Path,
    *,
    delimiter: str,
) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        writer.writerow(list(dataset.column_names))
        for row in dataset.to_rows():
            writer.writerow([row[name] for name in dataset.column_names])
    return output_path


def save_tabular_dataset_csv(dataset: TabularDataset, path: str | Path) -> Path:
    return _save_tabular_dataset_delimited(dataset, path, delimiter=",")


def save_tabular_dataset_tsv(dataset: TabularDataset, path: str | Path) -> Path:
    return _save_tabular_dataset_delimited(dataset, path, delimiter="\t")


def save_tabular_dataset_json(dataset: TabularDataset, path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(dataset.to_rows(), indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return output_path


def save_tabular_dataset_parquet(dataset: TabularDataset, path: str | Path) -> Path:
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Parquet support requires pyarrow. Install pyarrow in the active environment."
        ) from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    table = pa.table({name: values.tolist() for name, values in dataset.columns.items()})
    pq.write_table(table, output_path)
    return output_path


def save_tabular_dataset_xlsx(dataset: TabularDataset, path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "XLSX tabular support requires openpyxl. Install the 'files' extra."
        ) from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.append(list(dataset.column_names))
        for row in dataset.to_rows():
            worksheet.append([row[name] for name in dataset.column_names])
        workbook.save(output_path)
    finally:
        workbook.close()
    return output_path


def save_tabular_dataset(dataset: TabularDataset, path: str | Path) -> Path:
    output_path = Path(path)
    suffix = output_path.suffix.lower()
    if suffix == ".csv":
        return save_tabular_dataset_csv(dataset, output_path)
    if suffix == ".tsv":
        return save_tabular_dataset_tsv(dataset, output_path)
    if suffix == ".json":
        return save_tabular_dataset_json(dataset, output_path)
    if suffix == ".parquet":
        return save_tabular_dataset_parquet(dataset, output_path)
    if suffix == ".xlsx":
        return save_tabular_dataset_xlsx(dataset, output_path)
    supported = _supported_tabular_format_labels()
    raise ValueError(f"unsupported tabular format '{suffix or '<none>'}'; supported formats: {supported}")


__all__ = [
    "JoinSpec",
    "TabularColumn",
    "TabularDataset",
    "TabularSchema",
    "TabularSource",
    "TabularValidationReport",
    "load_tabular_dataset",
    "load_tabular_dataset_csv",
    "load_tabular_dataset_json",
    "load_tabular_dataset_parquet",
    "load_tabular_dataset_tsv",
    "load_tabular_dataset_xlsx",
    "parquet_support_is_available",
    "save_tabular_dataset",
    "save_tabular_dataset_csv",
    "save_tabular_dataset_json",
    "save_tabular_dataset_parquet",
    "save_tabular_dataset_tsv",
    "save_tabular_dataset_xlsx",
    "supported_tabular_formats",
    "xlsx_tabular_support_is_available",
]
