from __future__ import annotations

import csv
from dataclasses import dataclass, replace
import json
from pathlib import Path
from typing import Mapping, Sequence

import numpy as np
import torch

from spectral_packet_engine.runtime_files import atomic_output_path
from spectral_packet_engine.tabular import TabularDataset


_POSITION_PREFIXES = ("x=", "position=", "x:", "position:")
_DELIMITERS = {
    ".csv": ",",
    ".tsv": "\t",
}


def _parse_position_token(token: str) -> float:
    raw = token.strip()
    lowered = raw.lower()
    for prefix in _POSITION_PREFIXES:
        if lowered.startswith(prefix):
            raw = raw[len(prefix):].strip()
            break
    try:
        return float(raw)
    except ValueError as exc:
        raise ValueError(
            "profile table position headers must be numeric, or prefixed with 'x=' or 'position='"
        ) from exc


def excel_support_is_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ModuleNotFoundError:
        return False
    return True


def supported_profile_table_formats() -> dict[str, bool]:
    return {
        "csv": True,
        "tsv": True,
        "json": True,
        "xlsx": excel_support_is_available(),
    }


def _supported_profile_format_labels() -> str:
    supported = supported_profile_table_formats()
    available = [f".{name}" for name, enabled in supported.items() if enabled]
    unavailable = [f".{name}" for name, enabled in supported.items() if not enabled]
    if not unavailable:
        return ", ".join(available)
    return f"{', '.join(available)} (optional: {', '.join(unavailable)})"


@dataclass(frozen=True, slots=True)
class ProfileTableMaterializationConfig:
    time_column: str = "time"
    position_columns: tuple[str, ...] | None = None
    sort_by_time: bool = False
    source: str | None = None

    def __post_init__(self) -> None:
        time_column = str(self.time_column).strip()
        if not time_column:
            raise ValueError("time_column must not be empty")

        position_columns = None
        if self.position_columns is not None:
            position_columns = tuple(str(name).strip() for name in self.position_columns)
            if not position_columns:
                raise ValueError("position_columns must not be empty when provided")
            if any(not name for name in position_columns):
                raise ValueError("position_columns must not contain empty names")
            lowered = [name.lower() for name in position_columns]
            if len(set(lowered)) != len(lowered):
                raise ValueError("position_columns must be unique")
            if time_column.lower() in lowered:
                raise ValueError("position_columns must not include the time column")

        source = None if self.source is None else str(self.source)
        object.__setattr__(self, "time_column", time_column)
        object.__setattr__(self, "position_columns", position_columns)
        object.__setattr__(self, "source", source)


@dataclass(frozen=True, slots=True)
class ProfileTableLayout:
    time_column: str
    position_columns: tuple[str, ...]
    position_grid: np.ndarray

    def __post_init__(self) -> None:
        time_column = str(self.time_column).strip()
        position_columns = tuple(str(name).strip() for name in self.position_columns)
        position_grid = np.asarray(self.position_grid, dtype=np.float64)

        if not time_column:
            raise ValueError("time_column must not be empty")
        if not position_columns:
            raise ValueError("position_columns must not be empty")
        if any(not name for name in position_columns):
            raise ValueError("position_columns must not contain empty names")
        if len({name.lower() for name in position_columns}) != len(position_columns):
            raise ValueError("position_columns must be unique")
        if position_grid.ndim != 1:
            raise ValueError("position_grid must be one-dimensional")
        if position_grid.shape[0] != len(position_columns):
            raise ValueError("position_grid must align with position_columns")
        if not np.isfinite(position_grid).all():
            raise ValueError("position_grid must contain only finite values")
        if not np.all(np.diff(position_grid) > 0.0):
            raise ValueError("position_grid must be strictly increasing")

        object.__setattr__(self, "time_column", time_column)
        object.__setattr__(self, "position_columns", position_columns)
        object.__setattr__(self, "position_grid", position_grid)

    @property
    def num_positions(self) -> int:
        return int(self.position_grid.shape[0])


@dataclass(frozen=True, slots=True)
class ProfileTableLayoutResolution:
    time_column: str
    source_position_columns: tuple[str, ...]
    layout: ProfileTableLayout


def _coerce_materialization_config(
    *,
    config: ProfileTableMaterializationConfig | None = None,
    time_column: str = "time",
    position_columns: Sequence[str] | None = None,
    sort_by_time: bool = False,
    source: str | None = None,
) -> ProfileTableMaterializationConfig:
    if config is not None:
        if position_columns is not None or sort_by_time or source is not None or time_column != "time":
            raise ValueError("pass either config or explicit materialization arguments, not both")
        return config
    return ProfileTableMaterializationConfig(
        time_column=time_column,
        position_columns=None if position_columns is None else tuple(position_columns),
        sort_by_time=sort_by_time,
        source=source,
    )


def _resolve_column_name(columns: Mapping[str, np.ndarray], name: str) -> str:
    normalized = str(name).strip().lower()
    matches = [
        column_name
        for column_name in columns
        if column_name.strip().lower() == normalized
    ]
    if not matches:
        raise ValueError(f"tabular dataset does not contain a '{name}' column")
    if len(matches) > 1:
        raise ValueError(f"tabular dataset contains multiple '{name}' columns")
    return matches[0]


def _resolve_profile_table_layout(
    columns: Mapping[str, np.ndarray],
    *,
    config: ProfileTableMaterializationConfig,
) -> ProfileTableLayoutResolution:
    time_column = _resolve_column_name(columns, config.time_column)
    if config.position_columns is None:
        position_columns = [name for name in columns if name != time_column]
    else:
        position_columns = [_resolve_column_name(columns, name) for name in config.position_columns]
    if not position_columns:
        raise ValueError("tabular dataset must contain at least one position column")
    if len(set(position_columns)) != len(position_columns):
        raise ValueError("position_columns resolved to duplicate dataset columns")

    ordered = sorted(
        ((_parse_position_token(name), name) for name in position_columns),
        key=lambda item: item[0],
    )
    return ProfileTableLayoutResolution(
        time_column=time_column,
        source_position_columns=tuple(position_columns),
        layout=ProfileTableLayout(
            time_column=time_column,
            position_columns=tuple(name for _, name in ordered),
            position_grid=np.asarray([position for position, _ in ordered], dtype=np.float64),
        ),
    )


def _profile_table_from_columns(
    columns: Mapping[str, np.ndarray],
    *,
    layout: ProfileTableLayout,
    sort_by_time: bool = False,
    source: str | None = None,
) -> "ProfileTable":
    sample_times = np.asarray(columns[layout.time_column], dtype=np.float64)
    profiles = np.column_stack(
        [np.asarray(columns[name], dtype=np.float64) for name in layout.position_columns]
    )
    if sort_by_time and sample_times.shape[0] > 1:
        order = np.argsort(sample_times, kind="stable")
        sample_times = sample_times[order]
        profiles = profiles[order]
    return ProfileTable(
        position_grid=layout.position_grid,
        sample_times=sample_times,
        profiles=profiles,
        source=source,
    )


@dataclass(frozen=True, slots=True)
class ProfileTable:
    position_grid: np.ndarray
    sample_times: np.ndarray
    profiles: np.ndarray
    source: str | None = None

    def __post_init__(self) -> None:
        position_grid = np.asarray(self.position_grid, dtype=np.float64)
        sample_times = np.asarray(self.sample_times, dtype=np.float64)
        profiles = np.asarray(self.profiles, dtype=np.float64)
        source = None if self.source is None else str(self.source)

        if position_grid.ndim != 1:
            raise ValueError("position_grid must be one-dimensional")
        if sample_times.ndim != 1:
            raise ValueError("sample_times must be one-dimensional")
        if profiles.ndim != 2:
            raise ValueError("profiles must be two-dimensional")
        if profiles.shape != (sample_times.shape[0], position_grid.shape[0]):
            raise ValueError("profiles must have shape [sample, position]")
        if profiles.shape[0] < 1:
            raise ValueError("profiles must contain at least one sample")
        if position_grid.shape[0] < 2:
            raise ValueError("position_grid must contain at least two samples")
        if not np.isfinite(position_grid).all():
            raise ValueError("position_grid must contain only finite values")
        if not np.isfinite(sample_times).all():
            raise ValueError("sample_times must contain only finite values")
        if not np.isfinite(profiles).all():
            raise ValueError("profiles must contain only finite values")
        if not np.all(np.diff(position_grid) > 0.0):
            raise ValueError("position_grid must be strictly increasing")

        object.__setattr__(self, "position_grid", position_grid)
        object.__setattr__(self, "sample_times", sample_times)
        object.__setattr__(self, "profiles", profiles)
        object.__setattr__(self, "source", source)

    @property
    def num_samples(self) -> int:
        return int(self.profiles.shape[0])

    @property
    def num_positions(self) -> int:
        return int(self.position_grid.shape[0])

    def to_torch(
        self,
        *,
        dtype: torch.dtype = torch.float64,
        device: torch.device | str | None = None,
    ) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
        grid = torch.as_tensor(self.position_grid, dtype=dtype, device=device)
        times = torch.as_tensor(self.sample_times, dtype=dtype, device=device)
        profiles = torch.as_tensor(self.profiles, dtype=dtype, device=device)
        return grid, times, profiles

    def to_dict(self) -> dict[str, object]:
        return {
            "position_grid": self.position_grid.tolist(),
            "sample_times": self.sample_times.tolist(),
            "profiles": self.profiles.tolist(),
            "source": self.source,
        }


def profile_table_from_tabular_dataset(
    dataset: TabularDataset,
    *,
    config: ProfileTableMaterializationConfig | None = None,
    time_column: str = "time",
    position_columns: Sequence[str] | None = None,
    sort_by_time: bool = False,
    source: str | None = None,
) -> ProfileTable:
    default_source = None
    if dataset.source is not None and dataset.source.location is not None:
        default_source = dataset.source.location
    materialization = _coerce_materialization_config(
        config=config,
        time_column=time_column,
        position_columns=position_columns,
        sort_by_time=sort_by_time,
        source=source,
    )
    if materialization.source is None and default_source is not None:
        materialization = replace(materialization, source=default_source)
    resolution = _resolve_profile_table_layout(dataset.columns, config=materialization)
    return _profile_table_from_columns(
        dataset.columns,
        layout=resolution.layout,
        sort_by_time=materialization.sort_by_time,
        source=materialization.source,
    )


def resolve_profile_table_layout_from_tabular_dataset(
    dataset: TabularDataset,
    *,
    config: ProfileTableMaterializationConfig | None = None,
    time_column: str = "time",
    position_columns: Sequence[str] | None = None,
) -> ProfileTableLayoutResolution:
    materialization = _coerce_materialization_config(
        config=config,
        time_column=time_column,
        position_columns=position_columns,
    )
    return _resolve_profile_table_layout(dataset.columns, config=materialization)


def profile_table_layout_from_tabular_dataset(
    dataset: TabularDataset,
    *,
    config: ProfileTableMaterializationConfig | None = None,
    time_column: str = "time",
    position_columns: Sequence[str] | None = None,
) -> ProfileTableLayout:
    return resolve_profile_table_layout_from_tabular_dataset(
        dataset,
        config=config,
        time_column=time_column,
        position_columns=position_columns,
    ).layout


def tabular_dataset_from_profile_table(table: ProfileTable) -> TabularDataset:
    columns: dict[str, np.ndarray] = {
        "time": np.asarray(table.sample_times, dtype=np.float64),
    }
    for index, position in enumerate(table.position_grid.tolist()):
        columns[format(position, ".12g")] = np.asarray(table.profiles[:, index], dtype=np.float64)
    return TabularDataset(columns=columns)


def _needs_transpose(header: list[str], rows: list[list[str]], time_column: str) -> bool:
    """Detect a transposed layout where the first column holds numeric positions
    and the remaining headers are string profile names.  In the canonical layout
    the first header is *time_column* and the remaining headers are numeric
    positions.  When both:
      1. header[0] looks like a position label (e.g. 'x', 'position') and
      2. none of header[1:] parse as floats, but all values in column-0 do
    we conclude the table is transposed and should be flipped."""
    first = header[0].strip().lower()
    if first not in ("x", "position", "pos", "grid"):
        return False
    # Check that remaining headers are NOT numeric (i.e. they are profile names)
    for h in header[1:]:
        try:
            float(h.strip())
            return False  # at least one is numeric → not transposed
        except ValueError:
            continue
    # Check that first-column values ARE numeric (positions)
    for row in rows[:5]:
        if not row:
            continue
        try:
            float(row[0].strip())
        except ValueError:
            return False
    return True


def _transpose_rows(header: list[str], rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    """Transpose a table where rows are positions and columns are profiles
    into the canonical layout where rows are profiles and columns are positions."""
    profile_names = header[1:]  # string names become the "time" (sample) labels
    positions = [row[0] for row in rows]  # first column → position headers
    new_header = ["time"] + positions
    new_rows: list[list[str]] = []
    for col_idx, name in enumerate(profile_names):
        new_row = [name] + [row[col_idx + 1] for row in rows]
        new_rows.append(new_row)
    return new_header, new_rows


def _load_profile_table_delimited(
    path: str | Path,
    *,
    time_column: str = "time",
    delimiter: str,
) -> ProfileTable:
    input_path = Path(path)
    with input_path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise ValueError("profile table is empty") from exc

        if len(header) < 2:
            raise ValueError("profile table must contain a time column and at least one position column")

        raw_rows = [row for row in reader if row and not all(not cell.strip() for cell in row)]

    # Auto-transpose: first column is positions, remaining columns are profiles
    if _needs_transpose(header, raw_rows, time_column):
        header, raw_rows = _transpose_rows(header, raw_rows)

    if header[0].strip().lower() != time_column.lower():
        raise ValueError(f"profile table must start with a '{time_column}' column")

    position_grid = np.asarray([_parse_position_token(token) for token in header[1:]], dtype=np.float64)
    sample_times: list[float] = []
    profiles: list[list[float]] = []

    for row_index, row in enumerate(raw_rows, start=2):
        if len(row) != len(header):
            raise ValueError(
                f"profile table row {row_index} has {len(row)} columns, expected {len(header)}"
            )
        # time column: try float first, fall back to index for string labels
        try:
            sample_times.append(float(row[0]))
        except ValueError:
            sample_times.append(float(row_index - 2))
        profiles.append([float(cell) for cell in row[1:]])

    if not profiles:
        raise ValueError("profile table contains no data rows")

    return ProfileTable(
        position_grid=position_grid,
        sample_times=np.asarray(sample_times, dtype=np.float64),
        profiles=np.asarray(profiles, dtype=np.float64),
        source=str(input_path),
    )


def load_profile_table_csv(
    path: str | Path,
    *,
    time_column: str = "time",
) -> ProfileTable:
    return _load_profile_table_delimited(path, time_column=time_column, delimiter=",")


def load_profile_table_tsv(
    path: str | Path,
    *,
    time_column: str = "time",
) -> ProfileTable:
    return _load_profile_table_delimited(path, time_column=time_column, delimiter="\t")


def load_profile_table_json(
    path: str | Path,
) -> ProfileTable:
    input_path = Path(path)
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("profile JSON must contain a top-level object")

    try:
        position_grid = payload["position_grid"]
        sample_times = payload["sample_times"]
        profiles = payload["profiles"]
    except KeyError as exc:
        raise ValueError(
            "profile JSON must contain 'position_grid', 'sample_times', and 'profiles'"
        ) from exc

    return ProfileTable(
        position_grid=np.asarray(position_grid, dtype=np.float64),
        sample_times=np.asarray(sample_times, dtype=np.float64),
        profiles=np.asarray(profiles, dtype=np.float64),
        source=str(input_path),
    )


def load_profile_table_xlsx(
    path: str | Path,
    *,
    time_column: str = "time",
) -> ProfileTable:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "XLSX profile tables require openpyxl. Install the 'files' extra."
        ) from exc

    input_path = Path(path)
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("profile XLSX file is empty")

        header = ["" if cell is None else str(cell) for cell in rows[0]]
        if len(header) < 2:
            raise ValueError("profile XLSX file must contain a time column and at least one position column")
        if header[0].strip().lower() != time_column.lower():
            raise ValueError(f"profile XLSX file must start with a '{time_column}' column")

        position_grid = np.asarray([_parse_position_token(token) for token in header[1:]], dtype=np.float64)
        sample_times: list[float] = []
        profiles: list[list[float]] = []

        for row_index, row in enumerate(rows[1:], start=2):
            values = ["" if cell is None else str(cell) for cell in row]
            if not values or all(not value.strip() for value in values):
                continue
            if len(values) < len(header):
                values.extend([""] * (len(header) - len(values)))
            if len(values) != len(header):
                raise ValueError(
                    f"profile XLSX row {row_index} has {len(values)} columns, expected {len(header)}"
                )
            sample_times.append(float(values[0]))
            profiles.append([float(value) for value in values[1:]])

        if not profiles:
            raise ValueError("profile XLSX file contains no data rows")

        return ProfileTable(
            position_grid=position_grid,
            sample_times=np.asarray(sample_times, dtype=np.float64),
            profiles=np.asarray(profiles, dtype=np.float64),
            source=str(input_path),
        )
    finally:
        workbook.close()


def _save_profile_table_delimited(
    table: ProfileTable,
    path: str | Path,
    *,
    time_column: str = "time",
    position_format: str = ".12g",
    delimiter: str,
) -> Path:
    output_path = Path(path)
    header = [time_column, *[format(value, position_format) for value in table.position_grid]]
    with atomic_output_path(output_path) as temporary_path:
        with temporary_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle, delimiter=delimiter)
            writer.writerow(header)
            for sample_time, profile in zip(table.sample_times.tolist(), table.profiles.tolist()):
                writer.writerow([sample_time, *profile])
    return output_path


def save_profile_table_csv(
    table: ProfileTable,
    path: str | Path,
    *,
    time_column: str = "time",
    position_format: str = ".12g",
) -> Path:
    return _save_profile_table_delimited(
        table,
        path,
        time_column=time_column,
        position_format=position_format,
        delimiter=",",
    )


def save_profile_table_tsv(
    table: ProfileTable,
    path: str | Path,
    *,
    time_column: str = "time",
    position_format: str = ".12g",
) -> Path:
    return _save_profile_table_delimited(
        table,
        path,
        time_column=time_column,
        position_format=position_format,
        delimiter="\t",
    )


def save_profile_table_json(
    table: ProfileTable,
    path: str | Path,
) -> Path:
    output_path = Path(path)
    with atomic_output_path(output_path) as temporary_path:
        temporary_path.write_text(json.dumps(table.to_dict(), indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return output_path


def save_profile_table_xlsx(
    table: ProfileTable,
    path: str | Path,
    *,
    time_column: str = "time",
    position_format: str = ".12g",
) -> Path:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "XLSX profile tables require openpyxl. Install the 'files' extra."
        ) from exc

    output_path = Path(path)
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.append([time_column, *[format(value, position_format) for value in table.position_grid]])
        for sample_time, profile in zip(table.sample_times.tolist(), table.profiles.tolist()):
            worksheet.append([sample_time, *profile])
        with atomic_output_path(output_path) as temporary_path:
            workbook.save(temporary_path)
    finally:
        workbook.close()
    return output_path


def load_profile_table(path: str | Path) -> ProfileTable:
    input_path = Path(path)
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return load_profile_table_csv(input_path)
    if suffix == ".tsv":
        return load_profile_table_tsv(input_path)
    if suffix == ".json":
        return load_profile_table_json(input_path)
    if suffix == ".xlsx":
        return load_profile_table_xlsx(input_path)
    supported = _supported_profile_format_labels()
    raise ValueError(f"unsupported profile table format '{suffix or '<none>'}'; supported formats: {supported}")


def save_profile_table(table: ProfileTable, path: str | Path) -> Path:
    output_path = Path(path)
    suffix = output_path.suffix.lower()
    if suffix == ".csv":
        return save_profile_table_csv(table, output_path)
    if suffix == ".tsv":
        return save_profile_table_tsv(table, output_path)
    if suffix == ".json":
        return save_profile_table_json(table, output_path)
    if suffix == ".xlsx":
        return save_profile_table_xlsx(table, output_path)
    supported = _supported_profile_format_labels()
    raise ValueError(f"unsupported profile table format '{suffix or '<none>'}'; supported formats: {supported}")


__all__ = [
    "ProfileTable",
    "ProfileTableLayout",
    "ProfileTableLayoutResolution",
    "ProfileTableMaterializationConfig",
    "excel_support_is_available",
    "load_profile_table",
    "load_profile_table_csv",
    "load_profile_table_json",
    "load_profile_table_tsv",
    "load_profile_table_xlsx",
    "profile_table_layout_from_tabular_dataset",
    "profile_table_from_tabular_dataset",
    "resolve_profile_table_layout_from_tabular_dataset",
    "save_profile_table",
    "save_profile_table_csv",
    "save_profile_table_json",
    "save_profile_table_tsv",
    "save_profile_table_xlsx",
    "supported_profile_table_formats",
    "tabular_dataset_from_profile_table",
]
